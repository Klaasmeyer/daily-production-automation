#!/usr/bin/env python3
"""
Daily Production Automation
=============================
Reads all crew daily report files from the 'Daily's' subfolder,
updates the master production template, and writes a crew/subcontractor
contribution report.

Each run:
  - Syncs job rows in the template to match today's reports:
      * Removes job rows that have no report today
      * Adds new job rows for reports with jobs not yet in the template
  - Writes today's job-level footage into the correct rows
  - Writes contractor DAILY subtotal formulas (auto-sum of job rows)
  - Writes top-level daily summary formulas
  - Accumulates WEEKLY totals (resets each Monday)
  - Accumulates MONTHLY totals (resets on 1st of month)
  - Saves a dated copy AND updates the master template file

Usage:
    python daily_production_auto.py                   # uses today's date
    python daily_production_auto.py --date 03-24-2026 # override date (MM-DD-YYYY)
    python daily_production_auto.py --dry-run          # preview only, no files written
"""

import sys
import re
import argparse
from datetime import date, datetime
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  Run:  pip install openpyxl")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR           = Path(r"c:\Users\JoshuaCollver\OneDrive\Working Folder for Claude.AI\Daily Production")
DAILY_DIR          = BASE_DIR / "Daily's"
PRODUCTION_DIR     = BASE_DIR / "Production Report"
CREW_REPORT_DIR    = BASE_DIR / "Crew Report"
TEMPLATE_FILE      = PRODUCTION_DIR / "DAILY PRODUCTION TEMPLATE (7).xlsx"
TEMPLATE_SHEET     = "DAILY PRODUCTION TEMPLATE"

# ── Work type → job row column (cols C–L) ─────────────────────────────────────
JOB_COL = {
    "Bore":          "C",
    "Rock Bore":     "D",
    "Bore Inc":      "E",
    "Bore Rock Inc": "F",
    "Trench":        "G",
    "Trench Rock":   "H",
    "Plow":          "I",
    "Aerial":        "J",
    "Cable":         "K",
    "Drops":         "L",
}
COMMENT_COL = "N"

# Work types read from crew reports (subset of JOB_COL — no Bore Inc variants)
WORK_TYPES = ["Bore", "Rock Bore", "Trench", "Trench Rock", "Plow", "Aerial", "Cable", "Drops"]

# ── Summary block (rows 3–12) ──────────────────────────────────────────────────
# Col A = label, Col B = daily LF, Col D = weekly LF, Col F = monthly LF
# Col C = daily $, Col E = weekly $, Col G = monthly $ (formula cells — left to Excel)
SUMMARY_ROWS = {
    "Bore":          3,
    "Rock Bore":     4,
    "Bore Inc":      5,
    "Bore Rock Inc": 6,
    "Trench":        7,
    "Trench Rock":   8,
    "Plow":          9,
    "Aerial":        10,
    "Cable":         11,
    "Drops":         12,
}
DATE_CELL = "A1"   # date of last run

# ── Contractor definitions ─────────────────────────────────────────────────────
# (section_type, job_number_prefix, pattern found in column A of template)
CONTRACTOR_DEFS = [
    ("LG",         "01", "AT&T LG"),
    ("NATCO",      "08", "NATCO"),
    ("SVA",        "01", "AT&T SVA"),
    ("WINDSTREAM", "02", "WINDSTREAM"),
    ("YELCOTT",    "15", "YELCOTT"),
    ("ARTEL",      "12", "ARTEL"),
    ("FECC",       "39", "FECC"),
    ("PIPELINE",   None, "SOUTHERN PIPELINE"),
    ("AECC",       "32", "AECC"),
]

# ── Regex ──────────────────────────────────────────────────────────────────────
JOB_RE        = re.compile(r"\d{2}[-/]\d{4}[-/]\d{2}")
JOB_NUMBER_RE = re.compile(r"^\d{2}-\d{4}-\d{2}$")

# ── Header patterns for daily report layout detection ─────────────────────────
HEADER_PATTERNS = [
    ("rock bore",    "Rock Bore"),
    ("rock trench",  "Rock Trench"),
    ("daily ttl",    "Daily Total"),
    ("daily total",  "Daily Total"),
    ("bore",         "Bore"),
    ("plow",         "Plow"),
    ("trench",       "Trench"),
    ("aerial",       "Aerial"),
    ("cable",        "Cable"),
    ("drops",        "Drops"),
    ("remarks",      "Remarks"),
    ("date",         "Date"),
]


# ═════════════════════════════════════════════════════════════════════════════
# Utility helpers
# ═════════════════════════════════════════════════════════════════════════════

def _to_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except (ValueError, AttributeError):
                pass
    return None


def _norm_job(v):
    if not v:
        return None
    m = JOB_RE.search(str(v).strip())
    if m:
        return re.sub(r"[/]", "-", m.group())
    return None


def _norm_crew(name):
    return str(name or "UNKNOWN").strip().upper()


# ═════════════════════════════════════════════════════════════════════════════
# Daily report reading
# ═════════════════════════════════════════════════════════════════════════════

def detect_layout(ws):
    def cv(addr):
        v = ws[addr].value
        return str(v).strip() if v is not None else None

    job_a = _norm_job(cv("K5"))
    job_b = _norm_job(cv("M5"))

    if job_a:
        job_number = job_a
        wo_number  = cv("K4")
        location   = cv("K7")
        crew       = cv("K8")
        supervisor = cv("B11")
    elif job_b:
        job_number = job_b
        wo_number  = cv("M4")
        location   = cv("M6")
        crew       = cv("M7")
        supervisor = cv("B11")
    else:
        job_number = None
        wo_number = location = crew = supervisor = None
        for r in range(4, 13):
            for c in range(1, 20):
                jn = _norm_job(ws.cell(row=r, column=c).value)
                if jn:
                    job_number = jn
                    break
            if job_number:
                break
        supervisor = cv("B11")

    header_row = None
    col_map    = {}
    for row in ws.iter_rows(min_row=10, max_row=20):
        row_headers = {}
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                low = cell.value.strip().lower()
                for key, wt in HEADER_PATTERNS:
                    if key in low:
                        row_headers[wt] = cell.column - 1
                        break
        if "Date" in row_headers and ("Bore" in row_headers or "Plow" in row_headers):
            header_row = row[0].row
            col_map    = row_headers
            break

    if header_row is None:
        header_row = 14
        col_map = {
            "Date": 0, "Plow": 1, "Bore": 2, "Rock Bore": 3,
            "Trench": 4, "Rock Trench": 5, "Aerial": 6, "Cable": 7,
            "Drops": 8, "Daily Total": 9, "Remarks": 10,
        }

    return {
        "job_number": job_number,
        "wo_number":  wo_number,
        "location":   location,
        "crew":       crew,
        "supervisor": supervisor,
        "header_row": header_row,
        "col_map":    col_map,
        "data_start": header_row + 1,
    }


def _parse_footage_from_notes(text: str) -> dict:
    """
    Extract footage values from a remark/notes string for cases where a
    supervisor described work in text without filling in the numeric cells.

    Patterns recognized (case-insensitive):
      - "NUMBER' KEYWORD"   e.g. "300' bore", "1,200' aerial"
      - "KEYWORD NUMBER'"   e.g. "bore 300'", "plowed 580'"
      - "KEYWORD=NUMBER'"   e.g. "bfov(1)(1.25)=580'"

    Incomplete work ("inc"/"incomplete") is skipped for bore/aerial/cable/drops.
    For trench and trench rock, incomplete footage IS counted per user preference.
    Returns a dict with the same keys as WORK_TYPES, defaulting to 0.
    """
    totals = {wt: 0.0 for wt in WORK_TYPES}

    # Normalize: remove commas in numbers, lowercase
    text = re.sub(r"(\d),(\d)", r"\1\2", text)   # 1,036 → 1036
    text_lc = text.lower()

    # Work types where incomplete ("inc") footage should still be counted
    COUNT_INCOMPLETE = {"Trench", "Trench Rock"}

    # Keyword → work type mapping (longer/more-specific patterns first)
    KW_MAP = [
        (r"rock\s*bore",      "Rock Bore"),
        (r"rock\s*trench",    "Trench Rock"),
        (r"rock\s*saw",       "Trench Rock"),   # "40' rock saw" = Trench Rock
        (r"\bbore[d]?\b",     "Bore"),
        (r"\btrench[ed]?\b",  "Trench"),
        (r"\bplow[ed]?\b",    "Plow"),
        (r"\baerial\b",       "Aerial"),
        (r"\bcable\b",        "Cable"),
        (r"\bfiber\b",        "Cable"),
        (r"\bco\d+\b",        "Cable"),   # co24, co96, co144 etc.
        (r"\bdrop[s]?\b",     "Drops"),
    ]

    # Pattern: NUMBER' KEYWORD  or  KEYWORD NUMBER'  or  KEYWORD=NUMBER'
    number_pat = r"(\d+(?:\.\d+)?)'?"

    # Words after a match that indicate removal/decommissioning rather than
    # new installation — skip those matches entirely for all work types.
    REMOVAL_RE = re.compile(r"\b(?:delash|wreck|remov|demo|pull\s+out|take\s+out)")

    for kw_re, wt in KW_MAP:
        skip_inc = wt not in COUNT_INCOMPLETE

        # "KEYWORD ... NUMBER'" pattern
        for m in re.finditer(rf"{kw_re}\s*[=:\s]\s*{number_pat}", text_lc):
            after = text_lc[m.end():m.end() + 30]
            if REMOVAL_RE.search(after):
                continue
            if skip_inc and re.search(r"\binc(omplete)?\b", after):
                continue
            totals[wt] += float(m.group(1))

        # "NUMBER' KEYWORD" pattern
        for m in re.finditer(rf"{number_pat}\s*'?\s*(?:of\s+)?{kw_re}", text_lc):
            after = text_lc[m.end():m.end() + 30]
            if REMOVAL_RE.search(after):
                continue
            if skip_inc and re.search(r"\binc(omplete)?\b", after):
                continue
            totals[wt] += float(m.group(1))

    return totals


def read_daily_report(filepath: Path, target_date: date) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    layout = detect_layout(ws)
    col    = layout["col_map"]

    report = {
        "filename":        filepath.name,
        "job_number":      layout["job_number"],
        "wo_number":       layout["wo_number"],
        "location":        layout["location"],
        "crew":            layout["crew"],
        "supervisor":      layout["supervisor"],
        "today_values":    {wt: 0.0 for wt in WORK_TYPES},
        "today_remarks":   [],
        "today_entries":   [],
        "all_entries":     [],
        "used_last_entry": False,
    }

    consecutive_empty = 0   # rows with no date AND no other data
    last_date = None        # carry forward when a supervisor omits the date on a continuation row

    for row_tuple in ws.iter_rows(min_row=layout["data_start"], values_only=True):
        date_val = row_tuple[col.get("Date", 0)] if col.get("Date", 0) < len(row_tuple) else None

        if date_val is None:
            # Check if the row has ANY meaningful numeric footage content
            # (ignore text-only note rows — those don't carry a date forward)
            has_footage = any(
                isinstance(v, (int, float)) and v != 0
                for idx, v in enumerate(row_tuple)
                if idx != col.get("Date", 0)
            )
            has_any_content = any(
                v is not None and v != 0 and str(v).strip()
                for v in row_tuple
            )
            if not has_any_content:
                consecutive_empty += 1
                if consecutive_empty >= 5:
                    break   # past the end of real data
                continue

            consecutive_empty = 0

            # Continuation row (numeric footage or note text) with no date —
            # attribute it to the same date as the previous row.  This covers:
            #   • Multi-crew rows where the supervisor only wrote the date once
            #   • Text-only note rows that describe work (e.g. "40' rock saw inc")
            if last_date is not None:
                date_val = last_date  # fall through with carried date
            else:
                continue  # no prior date to carry, skip
        else:
            consecutive_empty = 0

        row_date = _to_date(date_val)
        if row_date is None:
            continue

        last_date = row_date  # update carry-forward tracker

        def gc(key):
            idx = col.get(key)
            return _to_num(row_tuple[idx]) if idx is not None and idx < len(row_tuple) else 0.0

        remarks_idx = col.get("Remarks")
        remarks_val = (
            row_tuple[remarks_idx]
            if remarks_idx is not None and remarks_idx < len(row_tuple)
            else None
        )
        remarks_text = str(remarks_val).strip() if remarks_val else ""

        entry = {
            "date":        row_date,
            "Plow":        gc("Plow"),
            "Bore":        gc("Bore"),
            "Rock Bore":   gc("Rock Bore"),
            "Trench":      gc("Trench"),
            "Trench Rock": gc("Rock Trench") or gc("Trench Rock"),
            "Aerial":      gc("Aerial"),
            "Cable":       gc("Cable"),
            "Drops":       gc("Drops"),
            "total":       gc("Daily Total"),
            "remarks":     remarks_text,
        }

        # If all footage cells are zero but remarks contain footage descriptions,
        # extract values from the notes text.
        cell_total = sum(entry[wt] for wt in WORK_TYPES)
        if cell_total == 0 and remarks_text:
            note_vals = _parse_footage_from_notes(remarks_text)
            note_total = sum(note_vals.values())
            if note_total > 0:
                for wt in WORK_TYPES:
                    entry[wt] = note_vals[wt]
                entry["total"]   = note_total
                entry["remarks"] = f"[from notes] {remarks_text}"

        report["all_entries"].append(entry)

    target_entries = [e for e in report["all_entries"] if e["date"] == target_date]
    report["today_entries"] = target_entries
    # No last-entry fallback: only work explicitly dated to target_date is counted.

    for wt in WORK_TYPES:
        report["today_values"][wt] = sum(e[wt] for e in report["today_entries"])

    report["today_remarks"] = [e["remarks"] for e in report["today_entries"] if e["remarks"]]
    return report


# ═════════════════════════════════════════════════════════════════════════════
# Deduplication
# ═════════════════════════════════════════════════════════════════════════════

def deduplicate_reports(reports: list) -> tuple:
    best       = {}
    all_by_job = defaultdict(list)

    for r in reports:
        jn = r["job_number"]
        if not jn:
            best[id(r)] = r
            continue
        all_by_job[jn].append(r)

    duplicates = []
    for jn, group in all_by_job.items():
        if len(group) == 1:
            best[jn] = group[0]
        else:
            def last_date(r):
                return r["all_entries"][-1]["date"] if r["all_entries"] else date.min
            group.sort(key=last_date, reverse=True)
            best[jn] = group[0]
            duplicates.extend(group[1:])

    return list(best.values()), duplicates


# ═════════════════════════════════════════════════════════════════════════════
# Template structure discovery
# ═════════════════════════════════════════════════════════════════════════════

def discover_sections(ws) -> list:
    """
    Scan template column A for contractor headers, then find job rows and
    DAILY/WEEKLY/MONTHLY total rows within each section.
    Returns list of section dicts sorted by header_row.
    """
    max_row  = ws.max_row
    sections = []

    for row_num in range(1, max_row + 1):
        a_val = ws.cell(row=row_num, column=1).value
        if not a_val or not isinstance(a_val, str):
            continue
        a_upper = a_val.strip().upper()
        for ctype, prefix, pattern in CONTRACTOR_DEFS:
            if pattern.upper() in a_upper:
                sections.append({
                    "name":        a_val.strip(),
                    "type":        ctype,
                    "prefix":      prefix,
                    "header_row":  row_num,
                    "job_rows":    [],
                    "daily_row":   None,
                    "weekly_row":  None,
                    "monthly_row": None,
                })
                break

    sections.sort(key=lambda s: s["header_row"])

    for i, sec in enumerate(sections):
        start = sec["header_row"]
        end   = sections[i + 1]["header_row"] - 1 if i + 1 < len(sections) else max_row
        for row_num in range(start, end + 1):
            b_val = ws.cell(row=row_num, column=2).value
            if b_val is None:
                continue
            b_str = str(b_val).strip()
            if JOB_NUMBER_RE.match(b_str):
                sec["job_rows"].append(row_num)
            elif b_str.upper() == "DAILY":
                sec["daily_row"] = row_num
            elif b_str.upper() == "WEEKLY":
                sec["weekly_row"] = row_num
            elif b_str.upper() == "MONTHLY":
                sec["monthly_row"] = row_num

    return sections


def build_job_row_map(ws) -> dict:
    """Return {job_number: row_number} for every job row in the template."""
    mapping = {}
    for row in ws.iter_rows(min_col=2, max_col=2):
        cell = row[0]
        if cell.value and isinstance(cell.value, str):
            val = cell.value.strip()
            if JOB_NUMBER_RE.match(val):
                mapping[val] = cell.row
    return mapping


# ═════════════════════════════════════════════════════════════════════════════
# Contractor type for new jobs
# ═════════════════════════════════════════════════════════════════════════════

def get_contractor_type(report: dict) -> str | None:
    """
    Determine which contractor section a job belongs to.
    For 01-prefix jobs, check the filename for 'SVA' vs 'LG'.
    """
    filename = (report.get("filename") or "").upper()
    job_num  = report.get("job_number") or ""
    prefix   = job_num[:2] if len(job_num) >= 2 else ""

    if prefix == "01":
        return "SVA" if "SVA" in filename else "LG"

    return {
        "08": "NATCO",
        "02": "WINDSTREAM",
        "15": "YELCOTT",
        "12": "ARTEL",
        "39": "FECC",
        "32": "AECC",
    }.get(prefix)


# ═════════════════════════════════════════════════════════════════════════════
# Job sync  (add new rows / remove rows with no report)
# ═════════════════════════════════════════════════════════════════════════════

def _remove_job_row(ws, row_num):
    """Delete a job row, preserving the contractor label if it lives in col A."""
    a_val = ws.cell(row=row_num, column=1).value
    if a_val and isinstance(a_val, str):
        # Move contractor name to the row below before deleting
        ws.cell(row=row_num + 1, column=1, value=a_val)
    ws.delete_rows(row_num)


def _add_job_row(ws, section, job_number):
    """Insert a blank job row just before the section's DAILY total row."""
    insert_at = section["daily_row"]
    ws.insert_rows(insert_at)
    ws.cell(row=insert_at, column=2, value=job_number)
    # Do NOT write contractor name into col A here — the section header already
    # has it at section["header_row"].  Writing it again creates a duplicate that
    # confuses discover_sections on the next call.


def sync_jobs(ws, reports: list) -> tuple:
    """
    Sync template job rows with today's reports:
      - Removes rows whose job numbers are absent from today's reports
      - Adds rows for job numbers that appear in reports but not the template
    Returns (added, removed) lists of job numbers.
    """
    sections    = discover_sections(ws)
    job_row_map = build_job_row_map(ws)
    report_jobs = {r["job_number"]: r for r in reports if r["job_number"]}

    to_remove = [jn for jn in job_row_map if jn not in report_jobs]
    to_add    = [(jn, r) for jn, r in report_jobs.items() if jn not in job_row_map]

    removed = []
    added   = []

    # Remove from bottom to top so earlier row numbers stay valid
    rows_to_remove = sorted(
        [(job_row_map[jn], jn) for jn in to_remove],
        reverse=True,
    )
    for row_num, jn in rows_to_remove:
        _remove_job_row(ws, row_num)
        removed.append(jn)

    # Re-discover after removals
    if removed:
        sections = discover_sections(ws)

    # Add new jobs to their correct contractor section
    for job_num, report in to_add:
        ctype      = get_contractor_type(report)
        target_sec = next((s for s in sections if s["type"] == ctype), None)
        if target_sec and target_sec["daily_row"] is not None:
            _add_job_row(ws, target_sec, job_num)
            added.append(job_num)
            # Re-discover after each insertion to keep row numbers current
            sections = discover_sections(ws)
        else:
            print(f"  WARNING: No section found for job {job_num} (contractor type: {ctype})")

    return added, removed


# ═════════════════════════════════════════════════════════════════════════════
# Formula writers
# ═════════════════════════════════════════════════════════════════════════════

def write_formulas(ws, sections: list):
    """
    Write structural Excel formulas:
      1. Each contractor DAILY row: =SUM(job rows) for each work-type column
      2. Summary daily LF cells (col B, rows 3–12): =SUM(all contractor DAILY rows)
    Dollar-value cells (C, E, G in the summary block) are left as-is so Excel
    recalculates them from the existing rate formulas in the template.
    """
    # 1. Contractor DAILY rows
    for sec in sections:
        if not sec["job_rows"] or sec["daily_row"] is None:
            continue
        first_job = sec["job_rows"][0]
        last_job  = sec["job_rows"][-1]
        for col in JOB_COL.values():
            ws[f"{col}{sec['daily_row']}"] = f"=SUM({col}{first_job}:{col}{last_job})"

    # 2. Summary daily LF — col B, rows 3–12
    # Summary Bore (B3) = SUM of col C across all DAILY rows  (C = bore column in job rows)
    # Summary Rock Bore (B4) = SUM of col D across all DAILY rows, etc.
    daily_rows = [sec["daily_row"] for sec in sections if sec["daily_row"] is not None]
    if daily_rows:
        for wt, row_num in SUMMARY_ROWS.items():
            job_col = JOB_COL.get(wt)
            if not job_col:
                continue
            refs = ",".join(f"{job_col}{r}" for r in daily_rows)
            ws[f"B{row_num}"] = f"=SUM({refs})"


# ═════════════════════════════════════════════════════════════════════════════
# Weekly / Monthly accumulation
# ═════════════════════════════════════════════════════════════════════════════

def accumulate_periods(ws, daily_totals: dict, target_date: date):
    """
    Update weekly (col D) and monthly (col F) accumulated LF in summary block.
    Reads last-run date from A1 to detect week/month boundaries.
    """
    last_run = None
    raw = ws[DATE_CELL].value
    if raw:
        if isinstance(raw, datetime):
            last_run = raw.date()
        elif isinstance(raw, date):
            last_run = raw
        elif isinstance(raw, str):
            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
                try:
                    last_run = datetime.strptime(raw.strip(), fmt).date()
                    break
                except ValueError:
                    pass

    if last_run == target_date:
        print("  (accumulation skipped — already processed today)")
        return

    new_week  = True
    new_month = True
    if last_run:
        if last_run.isocalendar()[:2] == target_date.isocalendar()[:2]:
            new_week = False
        if last_run.month == target_date.month and last_run.year == target_date.year:
            new_month = False

    labels = []
    if new_week:  labels.append("new week")
    if new_month: labels.append("new month")
    if labels:
        print(f"  (period reset: {', '.join(labels)})")

    for wt in WORK_TYPES:
        row_num   = SUMMARY_ROWS.get(wt)
        today_val = daily_totals.get(wt, 0.0)
        if not row_num:
            continue

        # Weekly LF → col D
        if new_week:
            ws[f"D{row_num}"] = round(today_val) if today_val else 0
        else:
            existing = _to_num(ws[f"D{row_num}"].value)
            ws[f"D{row_num}"] = round(existing + today_val) if (existing + today_val) else 0

        # Monthly LF → col F
        if new_month:
            ws[f"F{row_num}"] = round(today_val) if today_val else 0
        else:
            existing = _to_num(ws[f"F{row_num}"].value)
            ws[f"F{row_num}"] = round(existing + today_val) if (existing + today_val) else 0


# ═════════════════════════════════════════════════════════════════════════════
# Main template update
# ═════════════════════════════════════════════════════════════════════════════

def update_template(reports: list, target_date: date,
                    output_path: Path, dry_run: bool) -> tuple:

    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    ws = wb[TEMPLATE_SHEET]

    # ── 1. Sync job rows (add new, remove stale) ──────────────────────────
    added, removed = sync_jobs(ws, reports)
    if added:
        print(f"  Added {len(added)} new job(s):    {', '.join(sorted(added))}")
    if removed:
        print(f"  Removed {len(removed)} old job(s): {', '.join(sorted(removed))}")

    # ── 2. Re-discover structure after sync ───────────────────────────────
    sections    = discover_sections(ws)
    job_row_map = build_job_row_map(ws)

    matched   = []
    unmatched = []

    # ── 3. Clear all job-row footage so stale values don't linger ─────────
    for row_num in job_row_map.values():
        for col in JOB_COL.values():
            ws[f"{col}{row_num}"] = None
        ws[f"{COMMENT_COL}{row_num}"] = None

    # ── 4. Write today's job-level footage ────────────────────────────────
    for report in reports:
        job_num = report["job_number"]
        if not job_num:
            unmatched.append(report)
            continue
        row_num = job_row_map.get(job_num)
        if row_num is None:
            unmatched.append(report)
            continue

        vals = report["today_values"]
        for wt, col in JOB_COL.items():
            v = vals.get(wt, 0.0)
            if v:
                ws[f"{col}{row_num}"] = round(v)

        remarks = " | ".join(report["today_remarks"])
        if remarks:
            ws[f"{COMMENT_COL}{row_num}"] = remarks

        matched.append(report)

    # ── 5. Compute overall daily totals ───────────────────────────────────
    daily_totals = {wt: 0.0 for wt in WORK_TYPES}
    for report in reports:
        for wt in WORK_TYPES:
            daily_totals[wt] += report["today_values"].get(wt, 0.0)

    # ── 6. Write structural formulas ──────────────────────────────────────
    write_formulas(ws, sections)

    # ── 7. Accumulate weekly / monthly ────────────────────────────────────
    accumulate_periods(ws, daily_totals, target_date)

    # ── 8. Stamp date ─────────────────────────────────────────────────────
    ws[DATE_CELL] = target_date.strftime("%m/%d/%Y")

    # ── 9. Save ───────────────────────────────────────────────────────────
    if not dry_run:
        wb.save(output_path)
        try:
            wb.save(TEMPLATE_FILE)
            print(f"  Master template updated: {TEMPLATE_FILE.name}")
        except PermissionError:
            print(f"  WARNING: Could not update master template (file may be open in Excel).")
            print(f"  Close Excel and re-run, or copy {output_path.name} over it manually.")

    return matched, unmatched


# ═════════════════════════════════════════════════════════════════════════════
# Crew contribution report
# ═════════════════════════════════════════════════════════════════════════════

def split_crews(crew_field: str) -> list:
    if not crew_field:
        return ["UNKNOWN"]
    parts  = re.split(r"[,/]+", crew_field)
    result = [p.strip().upper() for p in parts if p.strip()]
    return result if result else ["UNKNOWN"]


def generate_crew_report(reports: list, target_date: date,
                          output_path: Path, dry_run: bool):
    crew_totals  = defaultdict(lambda: defaultdict(float))
    crew_details = defaultdict(list)
    all_remarks  = []

    for report in reports:
        raw_crew   = _norm_crew(report["crew"])
        crews      = split_crews(raw_crew)
        supervisor = str(report["supervisor"] or "").strip()
        job_num    = report["job_number"] or "N/A"
        location   = str(report["location"] or "").strip()
        vals       = report["today_values"]
        total_lf   = sum(vals.get(wt, 0.0) for wt in WORK_TYPES)
        flag       = " *" if report["used_last_entry"] else ""

        for wt in WORK_TYPES:
            crew_totals[raw_crew][wt] += vals.get(wt, 0.0)

        crew_details[raw_crew].append({
            "job":        job_num,
            "location":   location,
            "supervisor": supervisor,
            "values":     vals,
            "total":      total_lf,
            "remarks":    " | ".join(report["today_remarks"]),
            "flag":       flag,
            "filename":   report["filename"],
            "sub_crews":  crews if len(crews) > 1 else [],
        })

        for remark in report["today_remarks"]:
            if remark:
                all_remarks.append({
                    "crew":       raw_crew,
                    "supervisor": supervisor,
                    "job":        job_num,
                    "location":   location,
                    "remarks":    remark,
                })

    cwb      = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F497D")
    alt_fill = PatternFill("solid", fgColor="DCE6F1")
    c_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_align  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def style_hdr(ws, row, n):
        for c in range(1, n + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = c_align; cell.border = thin

    def sc(cell, alt=False, align=None):
        cell.fill      = alt_fill if alt else PatternFill()
        cell.alignment = align or c_align
        cell.border    = thin

    date_label = target_date.strftime("%B %d, %Y")

    # ── Crew Summary sheet ────────────────────────────────────────────────
    s1 = cwb.active
    s1.title = "Crew Summary"
    s1.append([f"Daily Production  --  {date_label}"])
    s1.cell(row=1, column=1).font = Font(bold=True, size=13)
    s1.row_dimensions[1].height = 22

    hdrs1 = ["Crew / Company"] + WORK_TYPES + ["Total LF"]
    for c, h in enumerate(hdrs1, 1):
        s1.cell(row=2, column=c, value=h)
    style_hdr(s1, 2, len(hdrs1))

    grand = {wt: 0.0 for wt in WORK_TYPES}
    for r, (crew, totals) in enumerate(sorted(crew_totals.items()), 3):
        alt = r % 2 == 0
        sc(s1.cell(row=r, column=1, value=crew), alt, l_align)
        row_total = 0.0
        for c, wt in enumerate(WORK_TYPES, 2):
            v = totals.get(wt, 0.0)
            grand[wt] += v
            sc(s1.cell(row=r, column=c, value=round(v) if v else None), alt)
            row_total += v
        tot_cell = s1.cell(row=r, column=len(hdrs1), value=round(row_total) if row_total else None)
        sc(tot_cell, alt); tot_cell.font = Font(bold=True)

    gt_row = len(crew_totals) + 3
    sc(s1.cell(row=gt_row, column=1, value="GRAND TOTAL"), align=l_align)
    for c, wt in enumerate(WORK_TYPES, 2):
        sc(s1.cell(row=gt_row, column=c, value=round(grand[wt]) if grand[wt] else None))
    sc(s1.cell(row=gt_row, column=len(hdrs1), value=round(sum(grand.values()))))
    style_hdr(s1, gt_row, len(hdrs1))

    s1.column_dimensions["A"].width = 30
    for c in range(2, len(hdrs1) + 1):
        s1.column_dimensions[get_column_letter(c)].width = 11

    # ── Job Detail sheet ──────────────────────────────────────────────────
    s2 = cwb.create_sheet("Job Detail")
    s2.append([f"Daily Production  --  {date_label}"])
    s2.cell(row=1, column=1).font = Font(bold=True, size=13)
    s2.row_dimensions[1].height = 22

    hdrs2 = ["Crew / Company", "Sub-Crews", "Supervisor", "Job Number",
              "Location"] + WORK_TYPES + ["Total LF", "Notes / Remarks", "Source File"]
    for c, h in enumerate(hdrs2, 1):
        s2.cell(row=2, column=c, value=h)
    style_hdr(s2, 2, len(hdrs2))

    r = 3
    for crew in sorted(crew_details.keys()):
        for info in crew_details[crew]:
            alt  = r % 2 == 0
            vals = info["values"]
            sub  = ", ".join(info["sub_crews"]) if info["sub_crews"] else ""
            sc(s2.cell(row=r, column=1, value=crew + info["flag"]), alt, l_align)
            sc(s2.cell(row=r, column=2, value=sub),                 alt, l_align)
            sc(s2.cell(row=r, column=3, value=info["supervisor"]),  alt, l_align)
            sc(s2.cell(row=r, column=4, value=info["job"]),         alt)
            sc(s2.cell(row=r, column=5, value=info["location"]),    alt, l_align)
            for c, wt in enumerate(WORK_TYPES, 6):
                v = vals.get(wt, 0.0)
                sc(s2.cell(row=r, column=c, value=round(v) if v else None), alt)
            sc(s2.cell(row=r, column=6 + len(WORK_TYPES),
                       value=round(info["total"]) if info["total"] else None), alt)
            sc(s2.cell(row=r, column=7 + len(WORK_TYPES), value=info["remarks"]),  alt, l_align)
            sc(s2.cell(row=r, column=8 + len(WORK_TYPES), value=info["filename"]), alt, l_align)
            r += 1

    s2.column_dimensions["A"].width = 30
    s2.column_dimensions["B"].width = 28
    s2.column_dimensions["C"].width = 20
    s2.column_dimensions["D"].width = 14
    s2.column_dimensions["E"].width = 24
    for c in range(6, 6 + len(WORK_TYPES) + 1):
        s2.column_dimensions[get_column_letter(c)].width = 11
    s2.column_dimensions[get_column_letter(7 + len(WORK_TYPES))].width = 42
    s2.column_dimensions[get_column_letter(8 + len(WORK_TYPES))].width = 42

    # ── Notes Review sheet ────────────────────────────────────────────────
    s3 = cwb.create_sheet("Notes Review")
    s3.append([f"Field Notes / Remarks -- {date_label}  (review for crew attribution)"])
    s3.cell(row=1, column=1).font = Font(bold=True, size=13)
    s3.row_dimensions[1].height = 22

    hdrs3 = ["Crew / Company", "Supervisor", "Job Number", "Location", "Remarks / Notes"]
    for c, h in enumerate(hdrs3, 1):
        s3.cell(row=2, column=c, value=h)
    style_hdr(s3, 2, len(hdrs3))

    if all_remarks:
        for r2, item in enumerate(all_remarks, 3):
            alt = r2 % 2 == 0
            sc(s3.cell(row=r2, column=1, value=item["crew"]),       alt, l_align)
            sc(s3.cell(row=r2, column=2, value=item["supervisor"]), alt, l_align)
            sc(s3.cell(row=r2, column=3, value=item["job"]),        alt)
            sc(s3.cell(row=r2, column=4, value=item["location"]),   alt, l_align)
            sc(s3.cell(row=r2, column=5, value=item["remarks"]),    alt, l_align)
    else:
        s3.cell(row=3, column=1, value="No remarks recorded for this date.")

    s3.column_dimensions["A"].width = 30
    s3.column_dimensions["B"].width = 20
    s3.column_dimensions["C"].width = 14
    s3.column_dimensions["D"].width = 24
    s3.column_dimensions["E"].width = 65

    if not dry_run:
        cwb.save(output_path)

    return crew_totals


# ═════════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════════

def parse_args():
    p = argparse.ArgumentParser(description="Daily Production Automation")
    p.add_argument("--date",    help="Override date MM-DD-YYYY (default: today)")
    p.add_argument("--dry-run", action="store_true", help="Preview only, no files written")
    return p.parse_args()


def main():
    args    = parse_args()
    dry_run = args.dry_run

    if args.date:
        try:
            target_date = datetime.strptime(args.date, "%m-%d-%Y").date()
        except ValueError:
            print(f"ERROR: Invalid date format '{args.date}'. Use MM-DD-YYYY.")
            sys.exit(1)
    else:
        target_date = date.today()

    date_str        = target_date.strftime("%m-%d-%Y")
    output_template = PRODUCTION_DIR  / f"Daily Production {date_str}.xlsx"
    output_crew     = CREW_REPORT_DIR / f"Crew Report {date_str}.xlsx"

    # Look for reports in order of priority:
    #   1. Daily's/YYYYMMDD/   (date-specific subfolder)
    #   2. Daily's/            (general drop folder)
    #   3. BASE_DIR            (root folder — reports placed there directly)
    date_subfolder = DAILY_DIR / target_date.strftime("%Y%m%d")
    if date_subfolder.is_dir() and any(date_subfolder.glob("*.xlsx")):
        reports_dir = date_subfolder
    elif DAILY_DIR.is_dir() and any(DAILY_DIR.glob("*.xlsx")):
        reports_dir = DAILY_DIR
    else:
        reports_dir = BASE_DIR

    sep = "-" * 72
    print(f"\n{sep}")
    print(f"  Daily Production Automation  --  {target_date.strftime('%A, %B %d, %Y')}")
    if dry_run:
        print("  [DRY RUN -- no files will be saved]")
    print(sep)
    print(f"\nReading reports from: {reports_dir}\n")

    # ── Read all daily reports ─────────────────────────────────────────────
    raw_reports = []
    errors      = []

    # Only read .xlsx files that look like crew reports (skip templates, scripts, etc.)
    # "daily production template" — master template file
    # "daily production "         — dated output copies (e.g. "Daily Production 04-13-2026.xlsx")
    # "crew report"               — crew contribution output files
    SKIP_PATTERNS = ("daily production template", "daily production ", "crew report", "daily_production_auto")
    for xlsx_file in sorted(reports_dir.glob("*.xlsx")):
        if any(p in xlsx_file.name.lower() for p in SKIP_PATTERNS):
            continue
        try:
            raw_reports.append(read_daily_report(xlsx_file, target_date))
        except Exception as exc:
            errors.append((xlsx_file.name, str(exc)))
            print(f"  ERROR reading {xlsx_file.name}: {exc}")

    reports, duplicates = deduplicate_reports(raw_reports)

    print(f"  {'File':<50}  {'Job #':<15}  {'Crew':<22}  {'LF':>6}  {'Note'}")
    print(f"  {'-'*50}  {'-'*15}  {'-'*22}  {'-'*6}  {'-'*18}")
    for r in reports:
        job   = r["job_number"] or "NO JOB#"
        crew  = (r["crew"] or "N/A")[:22]
        total = sum(r["today_values"].values())
        note  = "[last entry used]" if r["used_last_entry"] else ""
        print(f"  {r['filename'][:50]:<50}  {job:<15}  {crew:<22}  {total:>6.0f}  {note}")

    if duplicates:
        print(f"\n  Duplicate files skipped:")
        for d in duplicates:
            print(f"    - {d['filename']}")

    print(f"\n  {len(raw_reports)} files read  |  {len(duplicates)} duplicates skipped  |  {len(errors)} errors")

    # ── Update master template ─────────────────────────────────────────────
    print(f"\n{sep}")
    print("  Syncing jobs and updating template...")
    matched, unmatched = update_template(reports, target_date, output_template, dry_run)
    print(f"  Matched {len(matched)} jobs  |  {len(unmatched)} unmatched")

    if unmatched:
        print("\n  Jobs NOT matched (no section found — add contractor section manually):")
        for r in unmatched:
            print(f"    X  {(r['job_number'] or 'NO JOB#'):<22}  {r['filename']}")

    if not dry_run:
        print(f"  Saved -> {output_template.name}")

    # ── Crew report ────────────────────────────────────────────────────────
    print(f"\n{sep}")
    print("  Generating crew contribution report...")
    crew_totals = generate_crew_report(reports, target_date, output_crew, dry_run)

    print(f"\n  {'Crew':<30}  {'Bore':>6}  {'R.Bore':>6}  {'Trench':>7}  "
          f"{'Plow':>5}  {'Aerial':>6}  {'Cable':>5}  {'Drops':>5}  {'TOTAL':>7}")
    print(f"  {'-'*30}  {'-'*6}  {'-'*6}  {'-'*7}  {'-'*5}  {'-'*6}  {'-'*5}  {'-'*5}  {'-'*7}")
    for crew in sorted(crew_totals.keys()):
        t     = crew_totals[crew]
        total = sum(t.values())
        if total:
            print(
                f"  {crew[:30]:<30}  {t['Bore']:>6.0f}  {t['Rock Bore']:>6.0f}  "
                f"{t['Trench']:>7.0f}  {t['Plow']:>5.0f}  {t['Aerial']:>6.0f}  "
                f"{t['Cable']:>5.0f}  {t['Drops']:>5.0f}  {total:>7.0f}"
            )

    if not dry_run:
        print(f"  Saved -> {output_crew.name}")

    print(f"\n{sep}")
    print("  Done." if not dry_run else "  Dry run complete. No files written.")
    print(f"{sep}\n")


if __name__ == "__main__":
    main()
